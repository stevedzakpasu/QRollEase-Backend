from io import BytesIO
from fastapi.responses import FileResponse
from fastapi_mail import FastMail, MessageSchema, MessageType
from pydantic import EmailStr
from sqlmodel import select
from typing import List
from fastapi import APIRouter, Depends, HTTPException, UploadFile, status
from sqlmodel import Session
from app.api.deps import get_current_active_staff, get_current_active_superuser, get_current_verified_staff, get_current_verified_user
from app.core.deps import get_session
from app.crud.crud_attendance import attendance
from app.crud.crud_student import student
from app.crud.crud_lecture import lecture
from app.models.attendance import Attendance
from app.models.lecture import Lecture
from app.models.student import Student
from app.models.user import User
from app.schemas.attendance import AttendanceUpdate, AttendanceCreate, StaffAttendanceRead, StudentAttendanceRead
from openpyxl import Workbook
from app.core.settings import settings
router = APIRouter()

@router.get("/attendances", response_model=List[StaffAttendanceRead], dependencies=[Depends(get_current_active_superuser)])
def get_attendances(
    *, 
    session: Session = Depends(get_session),
    offset: int = 0,
    limit: int = 100
    ):
    attendances = attendance.get_multiple(session=session, offset=offset, limit=limit)
    return attendances



@router.post("/attendances", response_model=StaffAttendanceRead, dependencies=[Depends(get_current_verified_staff)])
def staff_create_attendance(
    *,
    session: Session = Depends(get_session),
    attendance_in: AttendanceCreate,
):

    student_in = student.get_by_student_id(session=session, student_id=attendance_in.student_id)

    if not student_in:
         raise HTTPException(status_code=400, detail="Student not found")
        
    lecture_in = lecture.get_by_secret(session=session, lecture_secret=attendance_in.lecture_secret)    

    if not lecture_in:
        raise HTTPException(status_code=404, detail="Lecture not found")

    existing_attendance = attendance.get_by_student_and_lecture_secret(
        session=session,
        student_id=attendance_in.student_id,
        lecture_secret=attendance_in.lecture_secret
    )

    if existing_attendance:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Attendance already recorded"
        )
    attendance_in.lecture_id = lecture_in.id


    new_attendance = attendance.create(session=session, obj_in=attendance_in)
    return new_attendance





@router.post("/students-attendances", response_model=StudentAttendanceRead, dependencies=[Depends(get_current_verified_user)])
def student_create_attendance(
    *,
    session: Session = Depends(get_session),
    attendance_in: AttendanceCreate ,
    user: User = Depends(get_current_verified_user),
    

    ):

    student_in = session.exec(
        select(Student).where(Student.student_id == attendance_in.student_id)
        ).first()
    
    current_student = session.exec(
        select(Student).join(User).where(User.id == user.id)
        ).first()

    if current_student != student_in:
         raise HTTPException(status_code=400, detail="Student ID does not match yours")
        
    if not student_in:
        raise HTTPException(status_code=404, detail="Student not found")
    

    lecture = session.exec(
        select(Lecture).where(Lecture.lecture_secret == attendance_in.lecture_secret)
        ).first()
  
        
    if not lecture:
        raise HTTPException(status_code=404, detail="Lecture not found")
    

    existing_attendance = attendance.get_by_student_and_lecture_secret(session=session, student_id=attendance_in.student_id, lecture_secret=attendance_in.lecture_secret)
    
   
    if existing_attendance:
            raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Attendance already recorded"
        )
    attendance_in.lecture_id = lecture.id


    new_attendance = attendance.create(session=session, obj_in=attendance_in)
    return new_attendance




@router.get("/attendances/{lecture_secret}", response_model=List[StaffAttendanceRead], dependencies=[Depends(get_current_active_staff)])
def get_attendance(
    *,
    session: Session = Depends(get_session),
    lecture_secret: str
    ):

    db_attendance = session.exec((select(Attendance).where(Attendance.lecture_secret== lecture_secret))).all()
    if not db_attendance:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Attendance not found"
        )
    return db_attendance


# @router.get("/attendances/{lecture_id}", response_model=List[StaffAttendanceRead], dependencies=[Depends(get_current_active_staff)])
# def get_attendance(
#     *,
#     session: Session = Depends(get_session),
#     lecture_id: int
#     ):

#     db_attendance = session.exec((select(Attendance).where(Attendance.lecture_id== lecture_id))).all()
#     if not db_attendance:
#         raise HTTPException(
#             status_code=status.HTTP_404_NOT_FOUND,
#             detail="Attendance not found"
#         )
#     return db_attendance



@router.get("/my_attendance/", response_model=List[StudentAttendanceRead], dependencies=[Depends(get_current_verified_user)])
def get_my_attendance(
    *,
    session: Session = Depends(get_session),
    user: User = Depends(get_current_verified_user)
    ):
    current_student = session.exec(
        select(Student).join(User).where(User.id == user.id)
        ).first()
    

    db_attendance = session.exec((select(Attendance).where(Attendance.student_id == current_student.student_id))).all()
    if not db_attendance:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Attendance not found"
        )
    return db_attendance

# @router.get("/attendances/{student_id}", response_model=List[StaffAttendanceRead], dependencies=[Depends(get_current_active_superuser)])
# def get_individual_attendance(
#     *,
#     session: Session = Depends(get_session),
#     student_id: str
#     ):



#     attendances = session.exec(select(Attendance).where(Attendance.student_id == student_id)).all()
    

#     if not attendances:
#         raise HTTPException(
#             status_code=status.HTTP_404_NOT_FOUND,
#             detail="Attendances not found"
#         )
#     return attendances

# @router.put("/attendances", response_model=AttendanceRead, dependencies=[Depends(get_current_active_superuser)])
# def update_attendance(
#     *,
#     session: Session = Depends(get_session),
#     attendance_in: AttendanceUpdate,
#     lecture_secret: str
#     ):
#     updated_attendance = attendance.update(session=session, attendance_id=attendance_id, obj_in=attendance_in)
#     if not updated_attendance:
#         raise HTTPException(
#             status_code=status.HTTP_404_NOT_FOUND,
#             detail="Attendance not found"
#         )
#     return updated_attendance



@router.delete("/attendances", dependencies=[Depends(get_current_active_superuser)])
def delete_attendance(
    *,
    session: Session = Depends(get_session),
    attendance_in :AttendanceCreate
    ):
    deleted_attendance = attendance.remove(session=session, lecture_secret=attendance_in.lecture_secret, student_id=attendance_in.student_id)
    if not deleted_attendance:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Attendance not found"
        )
    return {"success": "Attendance deleted successfully"}

@router.get("/generate_and_send_excel/{lecture_secret}", dependencies=[Depends(get_current_active_staff)])
async def generate_and_send_excel(
    *,
    session: Session = Depends(get_session),
    lecture_secret: str,
    name: str = "default_filename",
    staff: User = Depends(get_current_active_staff)
):
    db_attendance = session.exec(
        (select(Attendance).where(Attendance.lecture_secret == lecture_secret))
    ).all()

    if not db_attendance:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Attendance not found"
        )

    # Create a new Excel workbook
    workbook = Workbook()
    sheet = workbook.active

    # Add title to the first row
    sheet.append([name])
    sheet.append(["Student ID"])  # Heading for student IDs

    # Write student IDs to the Excel sheet
    for attendance in db_attendance:
        sheet.append([attendance.student_id])

    # Save the workbook to a BytesIO object
    excel_bytes_io = BytesIO()
    workbook.save(excel_bytes_io)
    excel_bytes_io.seek(0)  # Move the pointer to the beginning

    # Create an UploadFile instance based on BytesIO content
    upload_file = UploadFile(
        filename=f"{name}.xlsx",
        file=excel_bytes_io,
    )

    # Create FastMail instance
    mail = FastMail(settings.CONF)

    # Create email message
    message = MessageSchema(
        subject="Attendance Sheet",
        recipients=[EmailStr(staff.email)],  # Send to staff email
        body="Attached is the attendance sheet.",
        subtype="html",  # Specify the subtype
        attachments=[upload_file]
    )

    try:
        # Send email with attached Excel sheet
        await mail.send_message(message)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to send attendance sheet via email"
        )

    # Return a success message
    return {"message": "Attendance sheet generated and sent via email"}







# @router.get("/generate_excel/{lecture_secret}", dependencies=[Depends(get_current_active_staff)])
# def generate_excel(
#     *,
#     session: Session = Depends(get_session),
#     lecture_secret: str,
#     name: str = "default_filename"
# ):
#     db_attendance = session.exec(
#         (select(Attendance).where(Attendance.lecture_secret == lecture_secret))
#     ).all()

#     if not db_attendance:
#         raise HTTPException(
#             status_code=status.HTTP_404_NOT_FOUND,
#             detail="Attendance not found"
#         )

#     # Create a new Excel workbook
#     workbook = Workbook()
#     sheet = workbook.active

#     # Add title to the first row
#     sheet.append([name])
#     sheet.append(["Student ID"])  # Heading for student IDs

#     # Write student IDs to the Excel sheet
#     for attendance in db_attendance:
#         sheet.append([attendance.student_id])

#     # Prepare a temporary file to save the workbook
#     temp_filename = f"{name}.xlsx"
#     workbook.save(temp_filename)

#     # Prepare headers for download
#     headers = {
#         "Content-Disposition": f"attachment; filename={temp_filename}"
#     }

#     # Return the Excel file as a downloadable response
#     return FileResponse(temp_filename, headers=headers, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# @router.get("/generate_excel/{lecture_secret}", dependencies=[Depends(get_current_active_staff)])
# def generate_excel(
#     *,
#     session: Session = Depends(get_session),
#     lecture_secret: str,
#     name: str = "default_filename"
# ):
#     db_attendance = session.exec(
#         (select(Attendance).where(Attendance.lecture_secret == lecture_secret))
#     ).all()

#     if not db_attendance:
#         raise HTTPException(
#             status_code=status.HTTP_404_NOT_FOUND,
#             detail="Attendance not found"
#         )

#     # Create a new Excel workbook
#     workbook = Workbook()
#     sheet = workbook.active

#     # Write student IDs to the Excel sheet
#     for attendance in db_attendance:
#         sheet.append([attendance.student_id])

#     # Save the workbook to a temporary file
#     with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
#         file_path = tmp_file.name
#         workbook.save(file_path)

#     # Prepare headers for download
#     headers = {
#         "Content-Disposition": f"attachment; filename={name}.xlsx"
#     }

#     # Return the Excel file as a downloadable response
#     return FileResponse(file_path, headers=headers, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")